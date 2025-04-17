# --- START OF FILE TenFin-main/dashboard.py ---

#!/usr/bin/env python3

import os
import re
import io # Needed for BytesIO
import shutil
import json
from pathlib import Path
from typing import List, Dict, Any, Optional

from fastapi import FastAPI, Request, Form, HTTPException, status
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse # Ensure StreamingResponse is imported
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook # Ensure Workbook is imported

# --- Attempt to import custom module ---
try:
    from filter_engine import run_filter, INDIAN_STATES
except ImportError:
    print("ERROR: Could not import 'filter_engine'.")
    INDIAN_STATES = ["Error: State List Unavailable"]
    def run_filter(*args, **kwargs): raise RuntimeError("filter_engine not available.")

# === CONFIGURATION ===
try:
    SCRIPT_DIR = Path(__file__).parent.resolve()
except NameError:
    SCRIPT_DIR = Path('.').resolve()
BASE_PATH = SCRIPT_DIR / "scraped_data"
FILTERED_PATH = BASE_PATH / "Filtered Tenders"
TEMPLATES_DIR = SCRIPT_DIR / "templates"
FILTERED_TENDERS_FILENAME = "Filtered_Tenders.json" # Expect JSON

# --- FastAPI App Setup ---
app = FastAPI(title="TenFin Tender Dashboard")

# --- Template Engine Setup ---
templates: Optional[Jinja2Templates] = None
if not TEMPLATES_DIR.is_dir(): print(f"ERROR: Templates directory not found: '{TEMPLATES_DIR}'")
else: templates = Jinja2Templates(directory=str(TEMPLATES_DIR))

# === HELPER FUNCTIONS ===
def _validate_subdir(subdir: str) -> Path:
    # (Keep the existing _validate_subdir function)
    if not subdir or ".." in subdir or subdir.startswith(("/", "\\")):
        print(f"Validation failed for subdir: '{subdir}' (Invalid characters or empty)")
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid subdirectory name.")
    try:
        full_path = FILTERED_PATH / subdir
        resolved_path = full_path.resolve(strict=False)
        if resolved_path.parent != FILTERED_PATH.resolve(strict=False):
             print(f"Validation failed for subdir: '{subdir}' (Path traversal attempt detected: {resolved_path})")
             raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Path traversal attempt detected.")
    except Exception as e:
        print(f"Error resolving path for subdir '{subdir}': {e}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Error processing path.")
    return resolved_path

# === API ENDPOINTS ===

@app.get("/", response_class=HTMLResponse)
async def homepage(request: Request):
    # (Keep existing homepage function)
    if not templates: raise HTTPException(status_code=status.HTTP_503_SERVICE_UNAVAILABLE, detail="Template engine not configured.")
    subdirs: List[str] = []
    if FILTERED_PATH.is_dir():
        try: subdirs = sorted([item.name for item in FILTERED_PATH.iterdir() if item.is_dir()])
        except OSError as e: print(f"ERROR: Could not list directories in '{FILTERED_PATH}': {e}")
    else: print(f"Warning: Filtered data directory not found: '{FILTERED_PATH}'.")
    return templates.TemplateResponse("index.html", {"request": request, "subdirs": subdirs})


@app.get("/view/{subdir}", response_class=HTMLResponse)
async def view_tenders(request: Request, subdir: str):
    # (Keep existing view_tenders function - reads JSON)
    if not templates: raise HTTPException(status_code=status.HTTP_503_SERVICE_UNAVAILABLE, detail="Template engine not configured.")
    tenders: List[Dict[str, Any]] = []
    try:
        subdir_path = _validate_subdir(subdir)
        file_path = subdir_path / FILTERED_TENDERS_FILENAME
        if not file_path.is_file(): raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"File '{FILTERED_TENDERS_FILENAME}' not found in '{subdir}'.")
        print(f"Reading tender data from: {file_path}")
        with open(file_path, "r", encoding="utf-8") as f: tenders = json.load(f)
        if not isinstance(tenders, list):
             print(f"ERROR: JSON data in {file_path} is not a list.")
             raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Invalid tender data format.")
    except FileNotFoundError: raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"Tender data file not found for '{subdir}'.")
    except json.JSONDecodeError: raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Error reading tender data file format.")
    except HTTPException as e: raise e
    except Exception as e:
        print(f"ERROR: Failed to process view for subdir '{subdir}': {e}"); import traceback; traceback.print_exc()
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Error loading or processing tender data.")
    return templates.TemplateResponse("view.html", {"request": request, "subdir": subdir, "tenders": tenders})


# --- Keep existing endpoint for single download (optional, but good to have) ---
@app.get("/download/{subdir}")
async def download_tender_excel(subdir: str):
    """Downloads a single tender set as an Excel (.xlsx) file by reading JSON."""
    # (Keep existing single download function - reads JSON)
    tenders: List[Dict[str, Any]] = []
    try:
        subdir_path = _validate_subdir(subdir)
        file_path = subdir_path / FILTERED_TENDERS_FILENAME
        if not file_path.is_file(): raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"File '{FILTERED_TENDERS_FILENAME}' not found.")
        with open(file_path, "r", encoding="utf-8") as f: tenders = json.load(f)
        if not isinstance(tenders, list): raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Invalid tender data format.")
    except Exception as e: raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Error preparing download: {e}")

    wb = Workbook(); ws = wb.active; ws.title = subdir[:31]
    headers = ["start_date", "end_date", "opening_date", "tender_id", "title", "department", "state"]
    if tenders and isinstance(tenders[0], dict): headers = list(tenders[0].keys()) # Use keys from first dict if available
    ws.append(headers)
    for tender in tenders: ws.append([tender.get(header, "N/A") for header in headers])

    excel_buffer = io.BytesIO()
    try: wb.save(excel_buffer); excel_buffer.seek(0)
    except Exception as e: raise HTTPException(status_code=500, detail="Failed to generate Excel content.")
    safe_subdir = re.sub(r'[^\w\-]+', '_', subdir)
    filename = f"{safe_subdir}_{FILTERED_TENDERS_FILENAME.replace('.json', '.xlsx')}"
    return StreamingResponse(excel_buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=\"{filename}\""})


@app.get("/run-filter", response_class=HTMLResponse)
async def run_filter_form(request: Request):
    # (Keep existing run_filter_form function)
    if not templates: raise HTTPException(status_code=status.HTTP_503_SERVICE_UNAVAILABLE, detail="Template engine not configured.")
    tender_files: List[str] = []
    if BASE_PATH.is_dir():
        try: tender_files = sorted([ f.name for f in BASE_PATH.iterdir() if f.is_file() and f.name.startswith("Final_Tender_List") and f.name.endswith(".txt") ])
        except OSError as e: print(f"ERROR: Could not list tender source files in '{BASE_PATH}': {e}")
    else: print(f"Warning: Base data directory not found: '{BASE_PATH}'")
    current_states = INDIAN_STATES if 'INDIAN_STATES' in globals() else ["Error: State List Unavailable"]
    return templates.TemplateResponse("run_filter.html", {"request": request, "tender_files": tender_files, "states": current_states})


@app.post("/run-filter", response_class=HTMLResponse)
async def run_filter_submit(request: Request, keywords: str = Form(""), regex: bool = Form(False), filter_name: str = Form(...), file: str = Form(...), state: str = Form(...), start_date: str = Form(...), end_date: str = Form(...)):
    # (Keep existing run_filter_submit function)
    if not templates: raise HTTPException(status_code=status.HTTP_503_SERVICE_UNAVAILABLE, detail="Template engine not configured.")
    if not filter_name or ".." in filter_name or filter_name.startswith(("/", "\\")) or "/" in filter_name or "\\" in filter_name :
         return templates.TemplateResponse("error.html", {"request": request, "error": "Invalid Filter Name."}, status_code=status.HTTP_400_BAD_REQUEST)
    source_file_path = (BASE_PATH / file).resolve()
    if not source_file_path.is_file() or BASE_PATH.resolve() not in source_file_path.parents:
         return templates.TemplateResponse("error.html", {"request": request, "error": f"Selected source file '{file}' is invalid or not found."}, status_code=status.HTTP_400_BAD_REQUEST)
    try:
        keyword_list = [kw.strip() for kw in keywords.split(",") if kw.strip()]
        if 'run_filter' not in globals() or not callable(run_filter): raise RuntimeError("Filter engine not available.")
        result_path_str = run_filter(base_folder=BASE_PATH, tender_filename=file, keywords=keyword_list, use_regex=regex, filter_name=filter_name, state=state, start_date=start_date, end_date=end_date)
        expected_subdir = f"{filter_name} Tenders"
        if not result_path_str or not Path(result_path_str).is_file() or not result_path_str.endswith(".json"):
             print(f"Warning: run_filter returned unexpected path: {result_path_str}")
        return templates.TemplateResponse("success.html", {"request": request, "subdir": expected_subdir, "result_path": str(result_path_str)})
    except Exception as e:
        print(f"ERROR: Unexpected error running filter engine for '{filter_name}': {type(e).__name__}: {e}"); import traceback; traceback.print_exc()
        return templates.TemplateResponse("error.html", {"request": request, "error": f"Error running filter: {type(e).__name__}"}, status_code=500)


@app.post("/delete/{subdir}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_tender_set(subdir: str):
    # (Keep existing single delete function)
    try:
        folder_to_delete = _validate_subdir(subdir)
        if not folder_to_delete.is_dir(): raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"Directory '{subdir}' not found.")
        print(f"Attempting to delete directory: {folder_to_delete}")
        shutil.rmtree(folder_to_delete)
        print(f"Successfully deleted directory: {folder_to_delete}")
    except HTTPException as e: raise e
    except OSError as e: raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Failed to delete directory: {e.strerror}")
    except Exception as e: raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Unexpected error deleting: {e}")
    return RedirectResponse(url=app.url_path_for("homepage"), status_code=status.HTTP_303_SEE_OTHER)


@app.post("/bulk-delete", status_code=status.HTTP_204_NO_CONTENT)
async def bulk_delete_tender_sets(selected_subdirs: List[str] = Form(...)):
    # (Keep existing bulk delete function)
    deleted_count = 0; errors = []
    print(f"Received request to bulk delete: {selected_subdirs}")
    if not selected_subdirs: return RedirectResponse(url=app.url_path_for("homepage"), status_code=status.HTTP_303_SEE_OTHER)
    for subdir in selected_subdirs:
        try:
            folder_to_delete = _validate_subdir(subdir)
            if folder_to_delete.is_dir():
                print(f"Attempting to bulk delete directory: {folder_to_delete}")
                shutil.rmtree(folder_to_delete)
                print(f"Successfully bulk deleted directory: {folder_to_delete}")
                deleted_count += 1
            else: errors.append(f"Directory '{subdir}' not found.")
        except HTTPException as e: errors.append(f"Invalid subdirectory '{subdir}': {e.detail}")
        except OSError as e: errors.append(f"Failed to delete '{subdir}': {e.strerror}")
        except Exception as e: errors.append(f"Unexpected error deleting '{subdir}'.")
    if errors: print(f"Bulk delete completed with errors: {errors}")
    else: print(f"Bulk delete completed successfully for {deleted_count} item(s).")
    return RedirectResponse(url=app.url_path_for("homepage"), status_code=status.HTTP_303_SEE_OTHER)


# --- NEW ENDPOINT FOR BULK DOWNLOAD ---
@app.post("/bulk-download")
async def bulk_download_tender_excel(selected_subdirs: List[str] = Form(...)):
    """Creates a single Excel file with multiple sheets for selected tender sets."""
    print(f"Received request to bulk download: {selected_subdirs}")
    if not selected_subdirs:
        # Should be prevented by JS, but handle anyway
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No tender sets selected for download.")

    wb = Workbook()
    wb.remove(wb.active) # Remove default sheet
    processed_sheets = 0
    errors = []

    # Define consistent headers for all sheets
    headers = ["start_date", "end_date", "opening_date", "tender_id", "title", "department", "state"]

    for subdir in selected_subdirs:
        tenders: List[Dict[str, Any]] = []
        try:
            subdir_path = _validate_subdir(subdir)
            file_path = subdir_path / FILTERED_TENDERS_FILENAME

            if not file_path.is_file():
                print(f"Warning: JSON file not found for '{subdir}' during bulk download, skipping.")
                errors.append(f"Data for '{subdir}' not found.")
                continue # Skip to the next selected subdir

            # Read and parse JSON for this subdir
            with open(file_path, "r", encoding="utf-8") as f:
                tenders = json.load(f)
            if not isinstance(tenders, list):
                print(f"Warning: Invalid JSON format (not a list) for '{subdir}', skipping.")
                errors.append(f"Invalid data format for '{subdir}'.")
                continue # Skip to the next selected subdir

            # Create a new sheet for this subdir
            # Ensure sheet title is valid (max 31 chars, no invalid chars)
            safe_sheet_title = re.sub(r'[\\/*?:\[\]]+', '_', subdir)[:31]
            ws = wb.create_sheet(title=safe_sheet_title)
            ws.append(headers) # Add headers to the new sheet

            # Add data rows
            for tender in tenders:
                 # Ensure we use the defined header order and handle missing keys
                 ws.append([tender.get(header, "N/A") for header in headers])

            processed_sheets += 1
            print(f"Added sheet for '{subdir}' to bulk download.")

        except HTTPException as e:
            # Catch validation errors specifically
             print(f"ERROR: Validation failed for '{subdir}' during bulk download: {e.detail}")
             errors.append(f"Skipped invalid item '{subdir}'.")
        except json.JSONDecodeError:
            print(f"ERROR: Invalid JSON content for '{subdir}', skipping.")
            errors.append(f"Error reading data for '{subdir}'.")
        except Exception as e:
            print(f"ERROR: Unexpected error processing '{subdir}' for bulk download: {e}")
            import traceback
            traceback.print_exc()
            errors.append(f"Error processing '{subdir}'.")

    # Check if any sheets were successfully added
    if processed_sheets == 0:
        print("Bulk download failed: No valid data could be processed.")
        # Maybe redirect with an error message? For now, raise HTTP exception.
        error_details = "; ".join(errors)
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Could not generate download. Errors: {error_details}")

    # --- Save workbook to buffer ---
    excel_buffer = io.BytesIO()
    try:
        wb.save(excel_buffer)
        excel_buffer.seek(0)
    except Exception as e:
         print(f"ERROR: Failed saving bulk workbook to buffer: {e}")
         raise HTTPException(status_code=500, detail="Failed to generate bulk Excel file content.")

    # --- Create filename ---
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Bulk_Tender_Download_{timestamp}.xlsx"

    # --- Return StreamingResponse ---
    print(f"Sending bulk download file '{filename}' with {processed_sheets} sheets.")
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""}
    )

# --- END OF FILE TenFin-main/dashboard.py ---
